# -*- coding: utf-8 -*-
"""สร้างไฟล์ Excel ข้อมูลอะไหล่และการเบิกเข้า-ออก"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random

def create_parts_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "อะไหล่"

    # Headers
    headers = ["รหัสอะไหล่", "ชื่ออะไหล่ (ไทย)", "ชื่ออะไหล่ (อังกฤษ)", "หมวดหมู่",
               "รุ่นรถที่ใช้ได้", "ยี่ห้อ", "ราคา (บาท)", "ราคาทุน (บาท)",
               "จำนวนขั้นต่ำ", "OEM/Aftermarket", "คุณภาพ/แบรนด์"]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="TH SarabunPSK", size=14, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # ข้อมูลอะไหล่จากรูป spreadsheet
    parts_data = [
        # เครื่องยนต์
        ["HND-E001", "ลูกสูบ", "Piston", "เครื่องยนต์", "Wave 110i, Wave 125i", "Honda", 850, 550, 5, "OEM", "Honda Genuine"],
        ["HND-E002", "แหวนลูกสูบ", "Piston Ring", "เครื่องยนต์", "Wave 110i, Wave 125i", "Honda", 350, 220, 5, "OEM", "Honda Genuine"],
        ["HND-E003", "ประเก็นฝาสูบ", "Head Gasket", "เครื่องยนต์", "Click 125i, Click 150i", "Honda", 180, 100, 3, "Aftermarket", "NPP"],
        ["HND-E004", "วาล์วไอดี", "Intake Valve", "เครื่องยนต์", "Wave 125i, MSX 125", "Honda", 280, 180, 3, "OEM", "Honda Genuine"],
        ["HND-E005", "วาล์วไอเสีย", "Exhaust Valve", "เครื่องยนต์", "Wave 125i, MSX 125", "Honda", 280, 180, 3, "OEM", "Honda Genuine"],
        ["HND-E006", "ก้านสูบ", "Connecting Rod", "เครื่องยนต์", "Wave 110i", "Honda", 650, 420, 2, "OEM", "Honda Genuine"],
        ["HND-E007", "ปะเก็นครัช", "Clutch Gasket", "เครื่องยนต์", "Wave 125i, CB150R", "Honda", 95, 55, 5, "OEM", "Honda Genuine"],
        ["HND-E008", "แกนสูบ", "Cylinder", "เครื่องยนต์", "PCX 150", "Honda", 2800, 1800, 1, "OEM", "Honda Genuine"],
        ["HND-E009", "ฝาสูบ", "Cylinder Head", "เครื่องยนต์", "Click 125i", "Honda", 3500, 2200, 1, "OEM", "Honda Genuine"],
        ["HND-E010", "ประเก็นเสื้อสูบ", "Cylinder Gasket", "เครื่องยนต์", "Wave 110i, Wave 125i", "Honda", 120, 70, 5, "OEM", "Honda Genuine"],
        # กรอง/น้ำมัน
        ["HND-F001", "ไส้กรองอากาศ", "Air Filter", "กรอง/น้ำมัน", "Wave 110i", "Honda", 120, 70, 5, "OEM", "Honda Genuine"],
        ["HND-F002", "ไส้กรองน้ำมันเครื่อง", "Oil Filter", "กรอง/น้ำมัน", "PCX 160, ADV 160", "Honda", 65, 35, 10, "OEM", "Honda Genuine"],
        ["HND-F003", "กรองอากาศ Click", "Air Filter Click", "กรอง/น้ำมัน", "Click 125i, Click 150i", "Honda", 130, 80, 5, "OEM", "Honda Genuine"],
        ["HND-F004", "ไส้กรองน้ำมัน CB", "Oil Filter CB", "กรอง/น้ำมัน", "CB300R, CB500F", "Honda", 85, 50, 5, "OEM", "Honda Genuine"],
        ["HND-F005", "กรองเบนซิน", "Fuel Filter", "กรอง/น้ำมัน", "Wave 125i, Click 150i", "Honda", 95, 55, 8, "OEM", "Honda Genuine"],
        # ระบบเบรก
        ["HND-B001", "ผ้าเบรกหน้า", "Front Brake Pad", "ระบบเบรก", "Wave 125i, Wave 110i", "Honda", 220, 130, 5, "OEM", "Honda Genuine"],
        ["HND-B002", "ผ้าเบรกหลัง", "Rear Brake Pad", "ระบบเบรก", "Wave 125i, Wave 110i", "Honda", 180, 100, 5, "OEM", "Honda Genuine"],
        ["HND-B003", "จานเบรกหน้า", "Front Brake Disc", "ระบบเบรก", "Click 150i, PCX 160", "Honda", 1200, 750, 2, "Aftermarket", "Brembo"],
        ["HND-B004", "น้ำมันเบรก DOT4", "Brake Fluid DOT4", "ระบบเบรก", "ทุกรุ่น", "Honda", 150, 85, 10, "OEM", "Honda Genuine"],
        ["HND-B005", "สายเบรกหน้า", "Front Brake Cable", "ระบบเบรก", "Wave 110i, Dream", "Honda", 120, 70, 5, "OEM", "Honda Genuine"],
        # ระบบส่งกำลัง
        ["HND-T001", "โซ่ขับ 428-120L", "Drive Chain 428-120L", "ระบบส่งกำลัง", "Wave 125i, CBR150R", "Honda", 450, 280, 3, "Aftermarket", "DID"],
        ["HND-T002", "สายพาน PCX", "Drive Belt PCX", "ระบบส่งกำลัง", "PCX 160", "Honda", 1800, 1100, 2, "OEM", "Honda Genuine"],
        ["HND-T003", "สเตอร์หน้า 14T", "Front Sprocket 14T", "ระบบส่งกำลัง", "Wave 125i, CB150R", "Honda", 180, 100, 3, "OEM", "Honda Genuine"],
        ["HND-T004", "สเตอร์หลัง 35T", "Rear Sprocket 35T", "ระบบส่งกำลัง", "Wave 125i, CB150R", "Honda", 250, 150, 3, "OEM", "Honda Genuine"],
        ["HND-T005", "แผ่นคลัช", "Clutch Plate", "ระบบส่งกำลัง", "Wave 125i, MSX 125", "Honda", 380, 230, 3, "OEM", "Honda Genuine"],
        ["HND-T006", "ลูกปืนล้อหน้า", "Front Wheel Bearing", "ระบบส่งกำลัง", "Wave 110i, Wave 125i", "Honda", 120, 70, 5, "OEM", "NTN"],
        ["HND-T007", "สายพาน Click", "Drive Belt Click", "ระบบส่งกำลัง", "Click 125i, Click 150i", "Honda", 1200, 750, 2, "OEM", "Honda Genuine"],
        # ระบบไฟฟ้า
        ["HND-L001", "หัวเทียน NGK CPR7EA-9", "Spark Plug NGK CPR7EA-9", "ระบบไฟฟ้า", "Wave 110i, Wave 125i, Click 125i", "Honda", 120, 65, 10, "OEM", "NGK"],
        ["HND-L002", "แบตเตอรี่ YTZ5S", "Battery YTZ5S", "ระบบไฟฟ้า", "Wave 125i, Click 125i", "Honda", 850, 550, 3, "Aftermarket", "YUASA"],
        ["HND-L003", "หลอดไฟหน้า LED", "LED Headlight Bulb", "ระบบไฟฟ้า", "PCX 160, ADV 160", "Honda", 450, 280, 5, "Aftermarket", "Philips"],
        ["HND-L004", "ไดชาร์จ", "Stator Coil", "ระบบไฟฟ้า", "Wave 125i", "Honda", 1800, 1100, 1, "OEM", "Honda Genuine"],
        ["HND-L005", "CDI", "CDI Unit", "ระบบไฟฟ้า", "Wave 110i, Dream", "Honda", 950, 600, 2, "OEM", "Honda Genuine"],
        ["HND-L006", "รีเลย์สตาร์ท", "Starter Relay", "ระบบไฟฟ้า", "Click 150i, PCX 160", "Honda", 280, 170, 3, "OEM", "Honda Genuine"],
        ["HND-L007", "สวิทช์กุญแจ", "Ignition Switch", "ระบบไฟฟ้า", "Wave 110i, Wave 125i", "Honda", 650, 400, 2, "OEM", "Honda Genuine"],
        # ช่วงล่าง
        ["HND-S001", "โช้คหลัง", "Rear Shock Absorber", "ช่วงล่าง", "Wave 110i, Wave 125i", "Honda", 950, 600, 2, "Aftermarket", "YSS"],
        ["HND-S002", "ลูกปืนคอ", "Steering Bearing", "ช่วงล่าง", "Wave 125i, Click 150i", "Honda", 180, 100, 3, "OEM", "NTN"],
        ["HND-S003", "บุชยางแขนสวิง", "Swing Arm Bush", "ช่วงล่าง", "Wave 110i, Wave 125i", "Honda", 85, 45, 5, "OEM", "Honda Genuine"],
        ["HND-S004", "น้ำมันโช้ค", "Fork Oil", "ช่วงล่าง", "CBR150R, CB300R", "Honda", 250, 150, 5, "OEM", "Honda Genuine"],
        # ยาง/ล้อ
        ["HND-W001", "ยางนอก 80/90-17", "Tire 80/90-17", "ยาง/ล้อ", "Wave 110i, Wave 125i", "Honda", 650, 400, 4, "Aftermarket", "IRC"],
        ["HND-W002", "ยางนอก 90/90-14", "Tire 90/90-14", "ยาง/ล้อ", "Click 125i, Click 150i", "Honda", 750, 480, 4, "Aftermarket", "Michelin"],
        ["HND-W003", "ยางใน 17 นิ้ว", "Inner Tube 17\"", "ยาง/ล้อ", "Wave 110i, Wave 125i", "Honda", 120, 70, 5, "OEM", "Honda Genuine"],
        ["HND-W004", "ซี่ลวดล้อหน้า", "Front Spoke", "ยาง/ล้อ", "Wave 110i, Dream", "Honda", 350, 200, 3, "OEM", "Honda Genuine"],
        ["HND-W005", "ยางนอก 100/80-14", "Tire 100/80-14", "ยาง/ล้อ", "PCX 160, ADV 160", "Honda", 980, 620, 3, "Aftermarket", "Michelin"],
        # ตัวถัง/พลาสติก
        ["HND-P001", "บังโคลนหน้า", "Front Fender", "ตัวถัง/พลาสติก", "Wave 110i", "Honda", 480, 300, 2, "Aftermarket", "NPP"],
        ["HND-P002", "ชุดสี (ฝาข้าง)", "Side Cover Set", "ตัวถัง/พลาสติก", "Wave 125i", "Honda", 850, 550, 2, "OEM", "Honda Genuine"],
        ["HND-P003", "เบาะนั่ง", "Seat", "ตัวถัง/พลาสติก", "Click 150i", "Honda", 1200, 750, 1, "OEM", "Honda Genuine"],
        ["HND-P004", "กระจกมองข้าง", "Side Mirror", "ตัวถัง/พลาสติก", "ทุกรุ่น", "Honda", 180, 100, 5, "Aftermarket", "NPP"],
        ["HND-P005", "แฮนด์จับ", "Handle Grip", "ตัวถัง/พลาสติก", "Wave 110i, Wave 125i", "Honda", 120, 70, 5, "OEM", "Honda Genuine"],
        # ระบบเชื้อเพลิง
        ["HND-U001", "หัวฉีด", "Fuel Injector", "ระบบเชื้อเพลิง", "Click 125i, PCX 160", "Honda", 1500, 950, 2, "OEM", "Honda Genuine"],
        ["HND-U002", "ปั๊มเบนซิน", "Fuel Pump", "ระบบเชื้อเพลิง", "Wave 125i, Click 150i", "Honda", 2200, 1400, 1, "OEM", "Honda Genuine"],
        ["HND-U003", "ก๊อกน้ำมัน", "Fuel Cock", "ระบบเชื้อเพลิง", "Dream, Super Cub", "Honda", 180, 100, 3, "OEM", "Honda Genuine"],
        # ท่อไอเสีย
        ["HND-X001", "ท่อไอเสีย", "Exhaust Pipe", "ท่อไอเสีย", "Wave 125i", "Honda", 1800, 1100, 1, "OEM", "Honda Genuine"],
        ["HND-X002", "ปะเก็นท่อไอเสีย", "Exhaust Gasket", "ท่อไอเสีย", "Wave 110i, Click 125i", "Honda", 65, 35, 5, "OEM", "Honda Genuine"],
        # น้ำมัน/สารหล่อลื่น
        ["HND-O001", "น้ำมันเครื่อง 10W-30 0.8L", "Engine Oil 10W-30 0.8L", "น้ำมัน/สารหล่อลื่น", "Wave 110i, Wave 125i", "Honda", 180, 110, 10, "OEM", "Honda Genuine"],
        ["HND-O002", "น้ำมันเครื่อง 10W-40 1L", "Engine Oil 10W-40 1L", "น้ำมัน/สารหล่อลื่น", "PCX 160, CB300R, CB500F", "Honda", 280, 170, 10, "OEM", "Honda Genuine"],
        ["HND-O003", "น้ำมันเกียร์ 80W-90", "Gear Oil 80W-90", "น้ำมัน/สารหล่อลื่น", "Click 125i, PCX 160", "Honda", 120, 70, 10, "OEM", "Honda Genuine"],
        ["HND-O004", "จาระบี", "Grease", "น้ำมัน/สารหล่อลื่น", "ทุกรุ่น", "Honda", 85, 45, 10, "OEM", "Honda Genuine"],
        # สายควบคุม
        ["HND-C001", "สายคันเร่ง", "Throttle Cable", "สายควบคุม", "Wave 125i", "Honda", 150, 85, 3, "OEM", "Honda Genuine"],
        ["HND-C002", "สายครัช", "Clutch Cable", "สายควบคุม", "Wave 125i, CBR150R", "Honda", 180, 100, 3, "OEM", "Honda Genuine"],
        ["HND-C003", "สายเบรกหลัง", "Rear Brake Cable", "สายควบคุม", "Wave 110i, Dream", "Honda", 120, 70, 3, "OEM", "Honda Genuine"],
        ["HND-C004", "สายไมล์", "Speedometer Cable", "สายควบคุม", "Wave 110i, Wave 125i", "Honda", 130, 75, 3, "OEM", "Honda Genuine"],
    ]

    # Category colors
    cat_colors = {
        "เครื่องยนต์": "FFE6E6", "กรอง/น้ำมัน": "FFF2CC", "ระบบเบรก": "FFD9E6",
        "ระบบส่งกำลัง": "E6F3FF", "ระบบไฟฟ้า": "FFFFCC", "ช่วงล่าง": "E6FFE6",
        "ยาง/ล้อ": "F2E6FF", "ตัวถัง/พลาสติก": "FFE6CC", "ระบบเชื้อเพลิง": "CCF2FF",
        "ท่อไอเสีย": "E6E6E6", "น้ำมัน/สารหล่อลื่น": "CCFFCC", "สายควบคุม": "FFCCCC",
    }

    data_font = Font(name="TH SarabunPSK", size=13)
    for row_idx, part in enumerate(parts_data, 2):
        cat = part[3]
        fill = PatternFill(start_color=cat_colors.get(cat, "FFFFFF"), end_color=cat_colors.get(cat, "FFFFFF"), fill_type="solid")
        for col_idx, val in enumerate(part, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.fill = fill
            if col_idx in (7, 8):
                cell.number_format = '#,##0.00'

    # Column widths
    widths = [15, 30, 30, 18, 35, 10, 15, 15, 12, 15, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ===== Sheet 2: Stock Move In =====
    ws_in = wb.create_sheet("รับเข้า (Stock In)")
    in_headers = ["เลขที่เอกสาร", "รหัสอะไหล่", "ชื่ออะไหล่", "จำนวน", "วันที่", "อ้างอิง", "หมายเหตุ"]
    in_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    for col, h in enumerate(in_headers, 1):
        cell = ws_in.cell(row=1, column=col, value=h)
        cell.fill = in_fill
        cell.font = Font(name="TH SarabunPSK", size=14, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    stock_in_data = []
    base_date = datetime(2026, 1, 5)
    for i, part in enumerate(parts_data):
        qty = random.choice([10, 15, 20, 25, 30, 50])
        date = base_date + timedelta(days=random.randint(0, 60))
        doc_no = f"IN/2026/{str(i+1).zfill(5)}"
        po_ref = f"PO-2026-{str(i+1).zfill(3)}"
        stock_in_data.append([doc_no, part[0], part[1], qty, date.strftime("%Y-%m-%d"), po_ref, f"รับเข้าสต็อก {part[1]}"])

    # เพิ่มรอบ 2
    for i in range(0, 30):
        part = parts_data[i]
        qty = random.choice([5, 10, 15, 20])
        date = base_date + timedelta(days=random.randint(60, 120))
        idx = len(parts_data) + i + 1
        doc_no = f"IN/2026/{str(idx).zfill(5)}"
        po_ref = f"PO-2026-{str(idx).zfill(3)}"
        stock_in_data.append([doc_no, part[0], part[1], qty, date.strftime("%Y-%m-%d"), po_ref, f"เติมสต็อก {part[1]}"])

    for row_idx, data in enumerate(stock_in_data, 2):
        for col_idx, val in enumerate(data, 1):
            cell = ws_in.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border

    in_widths = [18, 15, 30, 10, 15, 18, 30]
    for i, w in enumerate(in_widths, 1):
        ws_in.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ===== Sheet 3: Stock Move Out =====
    ws_out = wb.create_sheet("เบิกออก (Stock Out)")
    out_headers = ["เลขที่เอกสาร", "รหัสอะไหล่", "ชื่ออะไหล่", "จำนวน", "วันที่", "อ้างอิง", "หมายเหตุ"]
    out_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    for col, h in enumerate(out_headers, 1):
        cell = ws_out.cell(row=1, column=col, value=h)
        cell.fill = out_fill
        cell.font = Font(name="TH SarabunPSK", size=14, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    stock_out_data = []
    reasons = ["ซ่อมรถลูกค้า", "เปลี่ยนตามระยะ", "ซ่อมบำรุง", "ขายหน้าร้าน", "เคลม"]
    for i in range(50):
        part = random.choice(parts_data[:40])
        qty = random.choice([1, 2, 3, 4, 5])
        date = base_date + timedelta(days=random.randint(10, 120))
        doc_no = f"OUT/2026/{str(i+1).zfill(5)}"
        so_ref = f"SO-2026-{str(i+1).zfill(3)}"
        reason = random.choice(reasons)
        stock_out_data.append([doc_no, part[0], part[1], qty, date.strftime("%Y-%m-%d"), so_ref, f"{reason} - {part[1]}"])

    for row_idx, data in enumerate(stock_out_data, 2):
        for col_idx, val in enumerate(data, 1):
            cell = ws_out.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border

    for i, w in enumerate(in_widths, 1):
        ws_out.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    filepath = "warehouse_spare_parts_data.xlsx"
    wb.save(filepath)
    print(f"✅ สร้างไฟล์ Excel สำเร็จ: {filepath}")
    print(f"   - Sheet 'อะไหล่': {len(parts_data)} รายการ")
    print(f"   - Sheet 'รับเข้า': {len(stock_in_data)} รายการ")
    print(f"   - Sheet 'เบิกออก': {len(stock_out_data)} รายการ")
    return parts_data, stock_in_data, stock_out_data

if __name__ == "__main__":
    create_parts_excel()
