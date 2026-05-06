# -*- coding: utf-8 -*-
"""Import data from honda_parts_dataset_v2.xlsx into Odoo via XML-RPC"""
import xmlrpc.client
import openpyxl
import random
from datetime import datetime, timedelta

URL = "http://localhost:8044"
DB = "wearhousepart"
USERNAME = "admin"
PASSWORD = "admin"

def connect():
    common = xmlrpc.client.ServerProxy(f"{URL}/xmlrpc/2/common")
    uid = common.authenticate(DB, USERNAME, PASSWORD, {})
    if not uid:
        raise Exception("Cannot connect to Odoo")
    models = xmlrpc.client.ServerProxy(f"{URL}/xmlrpc/2/object")
    print(f"Connected to Odoo (uid={uid})")
    return uid, models

def delete_old_data(uid, models):
    """Delete existing data"""
    # Delete stock moves first (has FK to parts)
    move_ids = models.execute_kw(DB, uid, PASSWORD, 'spare.stock.move', 'search', [[]])
    if move_ids:
        models.execute_kw(DB, uid, PASSWORD, 'spare.stock.move', 'unlink', [move_ids])
        print(f"  Deleted {len(move_ids)} stock moves")

    # Delete parts
    part_ids = models.execute_kw(DB, uid, PASSWORD, 'spare.part', 'search', [[]])
    if part_ids:
        models.execute_kw(DB, uid, PASSWORD, 'spare.part', 'unlink', [part_ids])
        print(f"  Deleted {len(part_ids)} parts")

    print("  Old data deleted")

def get_maps(uid, models):
    cats = models.execute_kw(DB, uid, PASSWORD, 'spare.category', 'search_read', [[]], {'fields': ['id', 'name']})
    cat_map = {c['name']: c['id'] for c in cats}
    motos = models.execute_kw(DB, uid, PASSWORD, 'spare.motorcycle', 'search_read', [[]], {'fields': ['id', 'name']})
    moto_map = {m['name']: m['id'] for m in motos}
    return cat_map, moto_map

def find_moto_ids(moto_text, moto_map):
    """Find motorcycle IDs from text like 'Wave 110i / Wave 125i'"""
    if not moto_text:
        return []
    ids = []
    for key, mid in moto_map.items():
        # Check if any part of the moto_text matches a motorcycle name
        parts = [p.strip() for p in str(moto_text).split('/')]
        for p in parts:
            p = p.strip()
            if not p:
                continue
            # Try matching
            if p.lower() in key.lower() or key.lower().startswith('honda ' + p.lower()):
                if mid not in ids:
                    ids.append(mid)
                break
    return ids

def import_parts_from_excel(uid, models, cat_map, moto_map):
    wb = openpyxl.load_workbook('honda_parts_dataset_v2.xlsx')
    ws = wb['อะไหล่ Honda ไทย']

    created = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        code, name_th, name_en, category, motos, year, price, stock_status, brand, warranty, oem_type, note = row
        if not code:
            continue

        cat_id = cat_map.get(category)
        if not cat_id:
            print(f"  WARN: category not found: {category}")
            continue

        moto_ids = find_moto_ids(motos, moto_map)
        pt = 'oem' if oem_type == 'OEM' else 'aftermarket'

        vals = {
            'name': name_th,
            'name_en': name_en,
            'code': str(code),
            'category_id': cat_id,
            'price': float(price) if price else 0,
            'cost': round(float(price) * 0.6, 2) if price else 0,
            'min_qty': 3,
            'part_type': pt,
            'quality_brand': brand or '',
            'brand': 'Honda',
            'warranty_months': int(warranty) if warranty else 0,
            'note': str(note) if note else '',
        }
        if moto_ids:
            vals['motorcycle_ids'] = [(6, 0, moto_ids)]

        try:
            pid = models.execute_kw(DB, uid, PASSWORD, 'spare.part', 'create', [vals])
            created.append({'id': pid, 'name': name_th, 'code': code, 'stock_status': stock_status})
            print(f"  [{len(created)}] {code} - {name_th} (id={pid})")
        except Exception as e:
            print(f"  ERROR {code} {name_th}: {e}")

    return created

def create_stock_moves(uid, models, parts):
    """Create stock in/out moves for all parts - no negative stock"""
    base = datetime(2026, 1, 5)
    stock_tracker = {}  # part_id -> current stock

    # === Stock In ===
    print("\nCreating Stock IN moves...")
    in_count = 0
    for i, p in enumerate(parts):
        qty = random.choice([8, 10, 12, 15, 20, 25, 30])
        if p['stock_status'] == 'สั่งจอง':
            qty = random.choice([3, 5, 8])
        date = (base + timedelta(days=random.randint(0, 30))).strftime("%Y-%m-%d %H:%M:%S")
        vals = {
            'part_id': p['id'],
            'qty': qty,
            'move_type': 'in',
            'state': 'done',
            'date': date,
            'reference': f"PO-2026-{str(i+1).zfill(3)}",
            'note': f"รับเข้าสต็อกครั้งแรก {p['name']}",
        }
        try:
            models.execute_kw(DB, uid, PASSWORD, 'spare.stock.move', 'create', [vals])
            stock_tracker[p['id']] = stock_tracker.get(p['id'], 0) + qty
            in_count += 1
        except Exception as e:
            print(f"  ERR IN: {e}")

    # Second stock in for 60% of parts
    for i, p in enumerate(parts[:int(len(parts)*0.6)]):
        qty = random.choice([5, 8, 10, 15])
        date = (base + timedelta(days=random.randint(30, 90))).strftime("%Y-%m-%d %H:%M:%S")
        vals = {
            'part_id': p['id'],
            'qty': qty,
            'move_type': 'in',
            'state': 'done',
            'date': date,
            'reference': f"PO-2026-R{str(i+1).zfill(3)}",
            'note': f"เติมสต็อก {p['name']}",
        }
        try:
            models.execute_kw(DB, uid, PASSWORD, 'spare.stock.move', 'create', [vals])
            stock_tracker[p['id']] = stock_tracker.get(p['id'], 0) + qty
            in_count += 1
        except Exception as e:
            print(f"  ERR IN2: {e}")

    print(f"  Created {in_count} stock IN records")

    # === Stock Out (ไม่ให้ติดลบ) ===
    print("\nCreating Stock OUT moves...")
    reasons = ["ซ่อมรถลูกค้า", "เปลี่ยนตามระยะ", "ซ่อมบำรุง", "ขายหน้าร้าน", "เคลมประกัน"]
    out_count = 0
    for i in range(80):
        p = random.choice(parts)
        available = stock_tracker.get(p['id'], 0)
        if available <= 0:
            continue  # ข้ามถ้าไม่มีสต็อก
        # เบิกไม่เกินครึ่งหนึ่งของสต็อก หรือไม่เกิน 5
        max_out = min(available - 1, 5)  # เหลืออย่างน้อย 1
        if max_out <= 0:
            continue
        qty = random.randint(1, max_out)
        date = (base + timedelta(days=random.randint(5, 110))).strftime("%Y-%m-%d %H:%M:%S")
        reason = random.choice(reasons)
        vals = {
            'part_id': p['id'],
            'qty': qty,
            'move_type': 'out',
            'state': 'done',
            'date': date,
            'reference': f"SO-2026-{str(i+1).zfill(3)}",
            'note': f"{reason} - {p['name']}",
        }
        try:
            models.execute_kw(DB, uid, PASSWORD, 'spare.stock.move', 'create', [vals])
            stock_tracker[p['id']] = stock_tracker.get(p['id'], 0) - qty
            out_count += 1
        except Exception as e:
            print(f"  ERR OUT: {e}")

    print(f"  Created {out_count} stock OUT records")

if __name__ == "__main__":
    print("=" * 50)
    print("Import from honda_parts_dataset_v2.xlsx")
    print("=" * 50)

    uid, models = connect()

    print("\nDeleting old data...")
    delete_old_data(uid, models)

    cat_map, moto_map = get_maps(uid, models)
    print(f"\nCategories: {len(cat_map)}")
    print(f"Motorcycles: {len(moto_map)}")

    print("\nImporting parts...")
    parts = import_parts_from_excel(uid, models, cat_map, moto_map)
    print(f"\nTotal parts imported: {len(parts)}")

    print("\nCreating stock movements...")
    create_stock_moves(uid, models, parts)

    print("\n" + "=" * 50)
    print("DONE!")
    print("=" * 50)
